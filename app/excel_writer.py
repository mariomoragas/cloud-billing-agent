from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference, ScatterChart, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_FONT = "Verdana"
TITLE_FILL = "C74634"
SUBTITLE_FILL = "1B1B1B"
CARD_FILL = "FBF7F4"
CARD_BORDER = "E7D7CF"
ACCENT_FILL = "C74634"
ACCENT_SOFT_FILL = "E8B4AE"
ACCENT_DARK_FILL = "8E2F23"
TEXT_DARK = "1B1B1B"
TEXT_MUTED = "5F5B57"
TEXT_LIGHT = "FFFFFF"


def write_billing_report(
    output_path: Path,
    raw_df: pd.DataFrame,
    service_summary_df: pd.DataFrame,
    region_summary_df: pd.DataFrame,
    oci_mapping_df: pd.DataFrame,
    extra_summaries: dict[str, pd.DataFrame] | None = None,
    data_quality_df: pd.DataFrame | None = None,
    llm_report_df: pd.DataFrame | None = None,
    llm_migration_df: pd.DataFrame | None = None,
    llm_recommendations_df: pd.DataFrame | None = None,
    llm_confidence_df: pd.DataFrame | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        raw_df.to_excel(writer, sheet_name="Raw_Data", index=False)
        service_summary_df.to_excel(writer, sheet_name="Resumo_Servicos", index=False)
        region_summary_df.to_excel(writer, sheet_name="Resumo_Regioes", index=False)
        oci_mapping_df.to_excel(writer, sheet_name="Mapeamento_OCI", index=False)
        if data_quality_df is not None and not data_quality_df.empty:
            data_quality_df.to_excel(writer, sheet_name="Data_Quality", index=False)
        if llm_report_df is not None and not llm_report_df.empty:
            llm_report_df.to_excel(writer, sheet_name="LLM_Resumo", index=False)
        if llm_migration_df is not None and not llm_migration_df.empty:
            llm_migration_df.to_excel(writer, sheet_name="LLM_Migracao", index=False)
        if llm_recommendations_df is not None and not llm_recommendations_df.empty:
            llm_recommendations_df.to_excel(writer, sheet_name="LLM_Recomendacoes", index=False)
        if llm_confidence_df is not None and not llm_confidence_df.empty:
            llm_confidence_df.to_excel(writer, sheet_name="LLM_Confianca", index=False)
        if extra_summaries:
            for sheet_name, summary_df in extra_summaries.items():
                if summary_df.empty:
                    continue
                safe_name = _sheet_name(sheet_name)
                summary_df.to_excel(writer, sheet_name=safe_name, index=False)
        _build_pending_sheet(writer, oci_mapping_df)

    workbook = load_workbook(output_path)
    _format_headers(workbook)
    _autosize_columns(workbook)
    _create_charts_sheet(workbook)
    workbook.save(output_path)


def _sheet_name(value: str) -> str:
    sanitized = value.replace("/", "_").replace("\\", "_")
    return sanitized[:31]


def _build_pending_sheet(writer: pd.ExcelWriter, oci_mapping_df: pd.DataFrame) -> None:
    pending = oci_mapping_df[oci_mapping_df["oci_service"] == "REVIEW_REQUIRED"].copy()
    if pending.empty:
        pending = pd.DataFrame(
            [{"status": "Sem pendencias de mapeamento para revisao manual."}]
        )
    pending.to_excel(writer, sheet_name="Pendencias", index=False)


def _format_headers(workbook) -> None:
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                is_header = cell.row == 1
                cell.font = Font(name=DEFAULT_FONT, bold=is_header)
                if is_header:
                    cell.fill = PatternFill(fill_type="solid", fgColor=SUBTITLE_FILL)
                    cell.font = Font(name=DEFAULT_FONT, bold=True, color=TEXT_LIGHT)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(vertical="top")


def _autosize_columns(workbook) -> None:
    for worksheet in workbook.worksheets:
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(length + 2, 12), 40
            )


def _create_charts_sheet(workbook) -> None:
    if "Charts" in workbook.sheetnames:
        del workbook["Charts"]

    charts_sheet = workbook.create_sheet("Charts")
    _style_charts_canvas(charts_sheet)
    _write_kpi_cards(workbook, charts_sheet)
    _add_service_cost_chart(workbook, charts_sheet)
    _add_service_cost_share_chart(workbook, charts_sheet)
    _add_region_cost_chart(workbook, charts_sheet)
    _add_migration_complexity_chart(workbook, charts_sheet)
    _add_optional_chart(
        workbook,
        charts_sheet,
        sheet_name="linked_account_name",
        title="Top contas por custo",
        category_axis="Conta",
        anchor="A50",
    )
    _add_optional_chart(
        workbook,
        charts_sheet,
        sheet_name="usage_type",
        title="Top usage types por custo",
        category_axis="Usage type",
        anchor="J50",
    )


def _style_charts_canvas(charts_sheet) -> None:
    charts_sheet.sheet_view.showGridLines = False
    charts_sheet["A1"] = "ORACLE"
    charts_sheet["A1"].font = Font(name=DEFAULT_FONT, bold=True, size=12, color=TEXT_LIGHT)
    charts_sheet["A2"] = "Cloud Billing Conversion Dashboard"
    charts_sheet["A2"].font = Font(name=DEFAULT_FONT, bold=True, size=20, color=TEXT_LIGHT)
    charts_sheet["A3"] = "Resumo executivo para analise financeira, racionalizacao de servicos e conversao para OCI."
    charts_sheet["A3"].font = Font(name=DEFAULT_FONT, italic=True, size=10, color="E5E7EB")
    charts_sheet.merge_cells("A1:L1")
    charts_sheet.merge_cells("A2:L2")
    charts_sheet.merge_cells("A3:L3")

    for cell_ref in ("A1", "A2", "A3"):
        charts_sheet[cell_ref].alignment = Alignment(vertical="center")
    for row in (1, 2):
        for col in range(1, 13):
            charts_sheet.cell(row=row, column=col).fill = PatternFill(
                fill_type="solid",
                fgColor=TITLE_FILL,
            )
    for col in range(1, 13):
        charts_sheet.cell(row=3, column=col).fill = PatternFill(
            fill_type="solid",
            fgColor=SUBTITLE_FILL,
        )

    for column in range(1, 22):
        charts_sheet.column_dimensions[get_column_letter(column)].width = 16
    for row in range(1, 120):
        charts_sheet.row_dimensions[row].height = 22


def _write_kpi_cards(workbook, charts_sheet) -> None:
    service_sheet = workbook["Resumo_Servicos"]
    mapping_sheet = workbook["Mapeamento_OCI"]
    raw_sheet = workbook["Raw_Data"]

    total_cost = _sum_column(service_sheet, "D")
    service_count = max(service_sheet.max_row - 1, 0)
    raw_rows = max(raw_sheet.max_row - 1, 0)
    unmapped_count = _count_matching(mapping_sheet, "G", "REVIEW_REQUIRED")

    cards = [
        ("A6:C9", "Custo total", _format_currency(total_cost), ACCENT_FILL, TEXT_LIGHT),
        ("D6:F9", "Servicos distintos", f"{service_count:,}", CARD_FILL, TEXT_DARK),
        ("G6:I9", "Linhas analisadas", f"{raw_rows:,}", CARD_FILL, TEXT_DARK),
        ("J6:L9", "Nao mapeados", f"{unmapped_count:,}", ACCENT_DARK_FILL, TEXT_LIGHT),
    ]

    for cell_range, label, value, fill_color, font_color in cards:
        charts_sheet.merge_cells(cell_range)
        top_left = charts_sheet[cell_range.split(":")[0]]
        top_left.value = f"{label}\n{value}"
        top_left.font = Font(name=DEFAULT_FONT, bold=True, size=14, color=font_color)
        top_left.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        top_left.fill = PatternFill(fill_type="solid", fgColor=fill_color)


def _sum_column(worksheet, column_letter: str) -> float:
    total = 0.0
    for row in range(2, worksheet.max_row + 1):
        value = worksheet[f"{column_letter}{row}"].value
        if isinstance(value, (int, float)):
            total += float(value)
    return total


def _count_matching(worksheet, column_letter: str, expected: str) -> int:
    count = 0
    for row in range(2, worksheet.max_row + 1):
        value = worksheet[f"{column_letter}{row}"].value
        if str(value or "").strip() == expected:
            count += 1
    return count


def _format_currency(value: float) -> str:
    return f"USD {value:,.2f}"


def _add_service_cost_chart(workbook, charts_sheet) -> None:
    service_category_col = _resolve_service_category_column(workbook["Resumo_Servicos"])
    service_sheet = _create_ranked_sheet(
        workbook,
        title="Chart_Service_Cost",
        source_sheet="Resumo_Servicos",
        category_col=service_category_col,
        value_col=4,
        top_n=8,
    )
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Top servicos por custo"
    chart.y_axis.title = "Servico"
    chart.x_axis.title = "Custo total"
    chart.height = 11
    chart.width = 8.8
    chart.legend = None
    chart.varyColors = False
    chart.gapWidth = 45
    chart.x_axis.numFmt = '#,##0.00'

    max_row = service_sheet.max_row
    if max_row < 2:
        return

    data = Reference(service_sheet, min_col=2, min_row=1, max_row=max_row)
    categories = Reference(service_sheet, min_col=2, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.series[0].graphicalProperties.solidFill = ACCENT_FILL
    chart.series[0].graphicalProperties.line.solidFill = ACCENT_FILL
    charts_sheet.add_chart(chart, "A12")


def _add_region_cost_chart(workbook, charts_sheet) -> None:
    region_sheet = _create_ranked_sheet(
        workbook,
        title="Chart_Region_Cost",
        source_sheet="Resumo_Regioes",
        category_col=2,
        value_col=3,
        top_n=6,
    )
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Top regioes por custo"
    chart.y_axis.title = "Regiao"
    chart.x_axis.title = "Custo total"
    chart.height = 11
    chart.width = 8.8
    chart.legend = None
    chart.gapWidth = 45
    chart.x_axis.numFmt = '#,##0.00'

    max_row = region_sheet.max_row
    if max_row < 2:
        return

    data = Reference(region_sheet, min_col=2, min_row=1, max_row=max_row)
    categories = Reference(region_sheet, min_col=2, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.series[0].graphicalProperties.solidFill = ACCENT_SOFT_FILL
    chart.series[0].graphicalProperties.line.solidFill = ACCENT_SOFT_FILL
    charts_sheet.add_chart(chart, "J12")


def _add_service_cost_share_chart(workbook, charts_sheet) -> None:
    service_category_col = _resolve_service_category_column(workbook["Resumo_Servicos"])
    service_sheet = _create_ranked_sheet(
        workbook,
        title="Chart_Service_Share",
        source_sheet="Resumo_Servicos",
        category_col=service_category_col,
        value_col=4,
        top_n=5,
    )
    max_row = service_sheet.max_row
    if max_row < 2:
        return

    chart = DoughnutChart()
    chart.title = "Composicao do custo por servico"
    chart.height = 12
    chart.width = 8.8
    chart.holeSize = 55
    chart.varyColors = True

    data = Reference(service_sheet, min_col=2, min_row=1, max_row=max_row)
    categories = Reference(service_sheet, min_col=2, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showCatName = True
    chart.dataLabels.showLeaderLines = True
    charts_sheet.add_chart(chart, "A31")


def _add_optional_chart(
    workbook,
    charts_sheet,
    sheet_name: str,
    title: str,
    category_axis: str,
    anchor: str,
) -> None:
    if sheet_name not in workbook.sheetnames:
        return

    data_sheet = _create_ranked_sheet(
        workbook,
        title=f"Chart_{sheet_name}",
        source_sheet=sheet_name,
        category_col=1,
        value_col=3,
        top_n=8,
    )
    max_row = data_sheet.max_row
    if max_row < 2:
        return

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = category_axis
    chart.x_axis.title = category_axis
    chart.height = 11
    chart.width = 8.8
    chart.legend = None
    chart.gapWidth = 45

    data = Reference(data_sheet, min_col=2, min_row=1, max_row=max_row)
    categories = Reference(data_sheet, min_col=1, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.series[0].graphicalProperties.solidFill = "9CA3AF"
    chart.series[0].graphicalProperties.line.solidFill = "9CA3AF"
    charts_sheet.add_chart(chart, anchor)


def _add_migration_complexity_chart(workbook, charts_sheet) -> None:
    if "Mapeamento_OCI" not in workbook.sheetnames:
        return

    matrix_sheet = _create_complexity_sheet(workbook, title="Chart_Migration_Complexity", top_n=10)
    if matrix_sheet.max_row < 2:
        return

    bar = BarChart()
    bar.type = "bar"
    bar.style = 10
    bar.title = "Complexidade de migracao (Top servicos por custo)"
    bar.y_axis.title = "Servico"
    bar.x_axis.title = "Complexity Score (1=Rehost ... 5=Retain)"
    bar.height = 10
    bar.width = 8.8
    bar.legend = None
    bar.varyColors = True
    bar.gapWidth = 45

    max_row = matrix_sheet.max_row
    data = Reference(matrix_sheet, min_col=3, min_row=1, max_row=max_row)
    categories = Reference(matrix_sheet, min_col=1, min_row=2, max_row=max_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(categories)
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True
    charts_sheet.add_chart(bar, "A69")

    scatter = ScatterChart()
    scatter.title = "Matriz custo x complexidade"
    scatter.x_axis.title = "Ranking por custo (1 = maior custo)"
    scatter.y_axis.title = "Complexity Score"
    scatter.height = 10
    scatter.width = 8.8
    scatter.legend = None
    scatter.y_axis.scaling.min = 0
    scatter.y_axis.scaling.max = 5
    scatter.y_axis.majorUnit = 1

    xvalues = Reference(matrix_sheet, min_col=4, min_row=2, max_row=max_row)
    yvalues = Reference(matrix_sheet, min_col=3, min_row=2, max_row=max_row)
    series = Series(yvalues, xvalues, title_from_data=False)
    scatter.series.append(series)
    charts_sheet.add_chart(scatter, "J69")


def _create_ranked_sheet(
    workbook,
    title: str,
    source_sheet: str,
    category_col: int,
    value_col: int,
    top_n: int,
):
    if title in workbook.sheetnames:
        del workbook[title]

    source = workbook[source_sheet]
    rows: list[tuple[str, float]] = []
    for row in range(2, source.max_row + 1):
        category = str(source.cell(row=row, column=category_col).value or "").strip()
        if not category:
            category = "UNSPECIFIED"
        value = source.cell(row=row, column=value_col).value
        if isinstance(value, (int, float)):
            rows.append((category, float(value)))

    rows.sort(key=lambda item: item[1], reverse=True)
    top_rows = rows[:top_n]
    remaining = rows[top_n:]
    others_total = sum(value for _, value in remaining)
    if others_total > 0:
        top_rows.append(("Outros", others_total))

    worksheet = workbook.create_sheet(title)
    worksheet.sheet_state = "hidden"
    worksheet["A1"] = "Categoria"
    worksheet["B1"] = "Valor"
    for index, (category, value) in enumerate(top_rows, start=2):
        worksheet[f"A{index}"] = category
        worksheet[f"B{index}"] = value
    return worksheet


def _create_complexity_sheet(workbook, title: str, top_n: int):
    if title in workbook.sheetnames:
        del workbook[title]

    source = workbook["Mapeamento_OCI"]
    header_map = {
        str(source.cell(row=1, column=column).value or "").strip(): column
        for column in range(1, source.max_column + 1)
    }
    required_headers = ["service_name_original", "total_cost", "complexity_score"]
    if not all(header in header_map for header in required_headers):
        worksheet = workbook.create_sheet(title)
        worksheet.sheet_state = "hidden"
        worksheet["A1"] = "service"
        worksheet["B1"] = "total_cost"
        worksheet["C1"] = "complexity_score"
        worksheet["D1"] = "cost_rank"
        return worksheet

    service_col = header_map["service_name_original"]
    product_col = header_map.get("primary_product_code")
    cost_col = header_map["total_cost"]
    complexity_col = header_map["complexity_score"]
    oci_col = header_map.get("oci_service")

    rows: list[tuple[str, float, int]] = []
    for row in range(2, source.max_row + 1):
        oci_service = str(source.cell(row=row, column=oci_col).value or "").strip() if oci_col else ""
        if oci_service == "REVIEW_REQUIRED":
            continue
        service_name = str(source.cell(row=row, column=service_col).value or "").strip()
        product_code = str(source.cell(row=row, column=product_col).value or "").strip() if product_col else ""
        label = product_code if product_code else (service_name or "UNSPECIFIED")

        raw_cost = source.cell(row=row, column=cost_col).value
        raw_complexity = source.cell(row=row, column=complexity_col).value
        if not isinstance(raw_cost, (int, float)):
            continue
        try:
            complexity = int(float(raw_complexity))
        except (TypeError, ValueError):
            complexity = 2
        complexity = max(1, min(5, complexity))
        rows.append((label, float(raw_cost), complexity))

    rows.sort(key=lambda item: item[1], reverse=True)
    top_rows = rows[:top_n]

    worksheet = workbook.create_sheet(title)
    worksheet.sheet_state = "hidden"
    worksheet["A1"] = "service"
    worksheet["B1"] = "total_cost"
    worksheet["C1"] = "complexity_score"
    worksheet["D1"] = "cost_rank"

    for index, (label, cost, complexity) in enumerate(top_rows, start=2):
        worksheet.cell(row=index, column=1, value=label)
        worksheet.cell(row=index, column=2, value=cost)
        worksheet.cell(row=index, column=3, value=complexity)
        worksheet.cell(row=index, column=4, value=index - 1)
    return worksheet


def _resolve_service_category_column(service_sheet) -> int:
    preferred_headers = ["chart_group_label", "service_name_original"]
    header_map = {
        str(service_sheet.cell(row=1, column=column).value or "").strip(): column
        for column in range(1, service_sheet.max_column + 1)
    }
    for header in preferred_headers:
        if header in header_map:
            return header_map[header]
    return 2
